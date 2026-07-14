from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth import login, get_user_model
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncMonth
from datetime import timedelta, date
from .models import *
from .forms import (
    SignUpForm, BrandForm, ProductGroupForm, SupplierForm,
    CustomerForm, InvoiceForm, InvoiceDetailFormSet, InvoicePaymentForm
)
from .ProductForm import ProductForm
from shared.export_mixins import ExportListMixin
from shared.mixins import PermissionRequiredAnyMixin
from shared.decorators import audit_action, permission_required_any, user_can_export
from billing.audit import log_action
from .column_config import get_visible_columns, get_all_columns, validate_visible_columns, DEFAULT_VISIBLE_COLUMNS
from .brand_column_config import (
    get_brand_visible_columns, get_all_brand_columns,
    validate_brand_visible_columns, BRAND_DEFAULT_VISIBLE_COLUMNS
)
from .productgroup_column_config import (
    get_productgroup_visible_columns, get_all_productgroup_columns,
    validate_productgroup_visible_columns, PRODUCTGROUP_DEFAULT_VISIBLE_COLUMNS
)
from .supplier_column_config import (
    get_supplier_visible_columns, get_all_supplier_columns,
    validate_supplier_visible_columns, SUPPLIER_DEFAULT_VISIBLE_COLUMNS
)
from shared.column_export import export_visible_columns_excel, export_visible_columns_pdf
from .customer_column_config import (
    get_customer_visible_columns, get_all_customer_columns,
    validate_customer_visible_columns, CUSTOMER_DEFAULT_VISIBLE_COLUMNS
)
from .invoice_column_config import (
    get_invoice_visible_columns, get_all_invoice_columns,
    validate_invoice_visible_columns, INVOICE_DEFAULT_VISIBLE_COLUMNS
)
from decimal import Decimal


# === HOME / DASHBOARD ===
@login_required
@audit_action('VIEW_HOME')
def home(request):
    """Panel principal: tarjetas resumen, gráficos, accesos rápidos,
    actividad reciente y alertas del sistema."""
    User = get_user_model()
    today = timezone.localdate()
    week_ago = today - timedelta(days=7)

    # --- Tarjetas resumen ---
    total_products = Product.objects.count()
    total_categories = ProductGroup.objects.count()
    total_brands = Brand.objects.count()
    total_suppliers = Supplier.objects.count()
    total_customers = Customer.objects.count()
    total_users = User.objects.count()
    total_sales = Invoice.objects.count()
    total_income = Invoice.objects.filter(is_active=True).exclude(estado='anulada').aggregate(s=Sum('total'))['s'] or 0

    from purchasing.models import Purchase
    total_purchases = Purchase.objects.count()

    low_stock_qs = Product.objects.filter(is_active=True, stock__gt=0, stock__lte=10)
    out_of_stock_qs = Product.objects.filter(is_active=True, stock=0)
    low_stock_count = low_stock_qs.count()
    out_of_stock_count = out_of_stock_qs.count()
    pending_credit_total = Invoice.objects.filter(tipo_pago='credito', saldo__gt=0).exclude(estado='anulada').aggregate(s=Sum('saldo'))['s'] or 0
    pending_credit_count = Invoice.objects.filter(tipo_pago='credito', saldo__gt=0).exclude(estado='anulada').count()

    # --- Gráfico: ventas por mes (últimos 6 meses) ---
    months = []
    cursor_year, cursor_month = today.year, today.month
    for i in range(5, -1, -1):
        y, m = cursor_year, cursor_month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(date(y, m, 1))

    sales_by_month = (
        Invoice.objects
        .filter(invoice_date__date__gte=months[0])
        .exclude(estado='anulada')
        .annotate(month=TruncMonth('invoice_date'))
        .values('month')
        .annotate(total=Sum('total'), count=Count('id'))
    )
    sales_map = {row['month'].strftime('%Y-%m'): row for row in sales_by_month if row['month']}
    month_labels = [m.strftime('%b %Y') for m in months]
    month_totals = [float(sales_map.get(m.strftime('%Y-%m'), {}).get('total') or 0) for m in months]
    month_counts = [sales_map.get(m.strftime('%Y-%m'), {}).get('count') or 0 for m in months]

    # --- Gráfico: productos más vendidos ---
    top_products_qs = (
        InvoiceDetail.objects
        .values('product__name')
        .annotate(qty=Sum('quantity'))
        .order_by('-qty')[:5]
    )
    top_products_labels = [row['product__name'] for row in top_products_qs]
    top_products_data = [row['qty'] for row in top_products_qs]

    # --- Gráfico: distribución de categorías ---
    category_dist_qs = (
        ProductGroup.objects
        .annotate(num_products=Count('products'))
        .filter(num_products__gt=0)
        .order_by('-num_products')
    )
    category_labels = [g.name for g in category_dist_qs]
    category_data = [g.num_products for g in category_dist_qs]

    # --- Gráfico: estado del stock (sustituye a "Compras vs Ventas", aún no implementado) ---
    healthy_stock_count = Product.objects.filter(is_active=True, stock__gt=10).count()
    stock_status_labels = ['Stock saludable', 'Stock bajo', 'Agotado']
    stock_status_data = [healthy_stock_count, low_stock_count, out_of_stock_count]

    # --- Indicadores (anillos KPI del panel) ---
    active_products = Product.objects.filter(is_active=True).count()

    def _pct(part, whole):
        return round((part / whole) * 100) if whole else 0

    stock_health_pct = _pct(healthy_stock_count, total_products)
    active_catalog_pct = _pct(active_products, total_products)
    collected = (total_income or 0) - (pending_credit_total or 0)
    collection_pct = max(0, min(100, _pct(collected, total_income))) if total_income else 100

    # --- Podium: top 3 productos más vendidos (para tarjetas de ranking) ---
    top_products_podium = [
        {'name': row['product__name'], 'qty': row['qty']}
        for row in top_products_qs
    ][:3]

    # --- Actividad reciente ---
    recent_sales = Invoice.objects.select_related('customer').order_by('-invoice_date')[:5]
    recent_products = Product.objects.select_related('brand', 'group').order_by('-created_at')[:5]
    recent_users = User.objects.order_by('-date_joined')[:5]

    # --- Alertas del sistema ---
    sales_today = Invoice.objects.filter(invoice_date__date=today)
    sales_today_count = sales_today.count()
    sales_today_total = sales_today.aggregate(s=Sum('total'))['s'] or 0
    new_customers_week = Customer.objects.filter(created_at__date__gte=week_ago).count()

    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_brands': total_brands,
        'total_suppliers': total_suppliers,
        'total_customers': total_customers,
        'total_users': total_users,
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'total_income': total_income,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'pending_credit_total': pending_credit_total,
        'pending_credit_count': pending_credit_count,

        'month_labels': month_labels,
        'month_totals': month_totals,
        'month_counts': month_counts,
        'top_products_labels': top_products_labels,
        'top_products_data': top_products_data,
        'category_labels': category_labels,
        'category_data': category_data,
        'stock_status_labels': stock_status_labels,
        'stock_status_data': stock_status_data,

        'stock_health_pct': stock_health_pct,
        'active_catalog_pct': active_catalog_pct,
        'collection_pct': collection_pct,
        'collected': collected,
        'top_products_podium': top_products_podium,

        'recent_sales': recent_sales,
        'recent_products': recent_products,
        'recent_users': recent_users,

        'low_stock_products': low_stock_qs.order_by('stock')[:5],
        'out_of_stock_products': out_of_stock_qs[:5],
        'sales_today_count': sales_today_count,
        'sales_today_total': sales_today_total,
        'new_customers_week': new_customers_week,
    }
    return render(request, 'billing/home.html', context)


# === REGISTRO ===
class SignUpView(PermissionRequiredAnyMixin, CreateView):
    permissions_required = ['auth.add_user']
    form_class = SignUpForm
    template_name = 'registration/signup.html'
    success_url = reverse_lazy('billing:verify_panel_code')

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.is_active = False
        self.object.set_unusable_password()
        self.object.save()
        from billing.services import _send_panel_verification_code
        _send_panel_verification_code(self.object, request=self.request)
        verify_url = reverse_lazy('billing:verify_panel_code')
        from django.utils.html import format_html
        messages.success(
            self.request,
            format_html(
                'Cuenta creada. Revisa tu correo para el código de verificación. Tu usuario es: <strong>{}</strong>',
                self.object.username,
            )
        )
        return redirect(verify_url)

# === BRAND (FBV) ===
@permission_required_any('billing.view_brand')
@audit_action('LIST_BRANDS')
def brand_list(request):
    qs = Brand.objects.annotate(product_count=Count('products'))
    name = request.GET.get('name', '').strip()
    is_active = request.GET.get('is_active', '')
    if name:
        qs = qs.filter(name__icontains=name)
    if is_active in ('0', '1'):
        qs = qs.filter(is_active=is_active == '1')

    # Columnas visibles (selector de columnas)
    visible_columns_list = request.session.get('brand_visible_columns', BRAND_DEFAULT_VISIBLE_COLUMNS)
    is_valid, visible_columns_list = validate_brand_visible_columns(visible_columns_list)

    def get_export_value(obj, col_key):
        if col_key == 'name':
            return obj.name
        elif col_key == 'description':
            return obj.description or '-'
        elif col_key == 'product_count':
            return obj.product_count
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        return getattr(obj, col_key, '-')

    # Export (respeta las columnas visibles seleccionadas)
    export = request.GET.get('export')
    if export in ('excel', 'pdf') and not user_can_export(request, 'billing.export_brand'):
        messages.error(request, 'No tienes permiso para exportar esta información.')
        export = None
    if export in ('excel', 'pdf'):
        all_columns = get_all_brand_columns()
        if export == 'excel':
            return export_visible_columns_excel(qs, visible_columns_list, all_columns, get_export_value, 'Marcas', 'Marcas')
        return export_visible_columns_pdf(qs, visible_columns_list, all_columns, get_export_value, 'Listado de Marcas', 'Marcas')

    # Paginación manual
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    params.pop('page', None)

    return render(request, 'billing/brand_list.html', {
        'brands': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'search_params': params.urlencode(),
        'visible_columns': get_brand_visible_columns(visible_columns_list),
        'all_columns': get_all_brand_columns(),
        'visible_columns_list': visible_columns_list,
    })

@login_required
def brand_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de marcas"""
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        is_valid, validated_columns = validate_brand_visible_columns(visible_columns)
        request.session['brand_visible_columns'] = validated_columns
        request.session.modified = True
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_brand_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@permission_required_any('billing.add_brand')
@audit_action('CREATE_BRAND')
def brand_create(request):
    if request.method == 'POST':
        form = BrandForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca creada exitosamente!')
            return redirect('billing:brand_list')
    else:
        form = BrandForm()
    return render(request, 'billing/brand_form.html', {'form': form, 'title': 'Crear Marca'})

@permission_required_any('billing.view_brand')
@audit_action('VIEW_BRAND')
def brand_detail(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    products = brand.products.all()[:10]
    return render(request, 'billing/brand_detail.html', {
        'brand': brand,
        'products': products,
        'product_count': brand.products.count(),
    })

@permission_required_any('billing.change_brand')
@audit_action('UPDATE_BRAND')
def brand_update(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, request.FILES, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca actualizada exitosamente!')
            return redirect('billing:brand_list')
    else:
        form = BrandForm(instance=brand)
    return render(request, 'billing/brand_form.html', {'form': form, 'title': 'Editar Marca'})

@permission_required_any('billing.delete_brand')
@audit_action('DELETE_BRAND')
def brand_delete(request, pk):
    from django.db.models.deletion import ProtectedError
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        try:
            brand.delete()
            messages.success(request, 'Marca eliminada exitosamente!')
        except ProtectedError:
            messages.error(request, 'No se puede eliminar esta marca porque tiene productos asociados.')
        return redirect('billing:brand_list')
    return render(request, 'billing/brand_confirm_delete.html', {'object': brand})


# === PRODUCTGROUP (CBV) ===
class ProductGroupListView(PermissionRequiredAnyMixin, ExportListMixin, ListView):
    permissions_required = ['billing.view_productgroup']
    model = ProductGroup
    template_name = 'billing/product_group_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Grupos de Productos'
    export_fields = [
        ('Nombre', 'name'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
        ('Creación', lambda o: o.created_at.strftime('%d/%m/%Y')),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf') and not user_can_export(request, 'billing.export_productgroup'):
            messages.error(request, 'No tienes permiso para exportar esta información.')
            fmt = None
        if fmt in ('excel', 'pdf'):
            visible_columns = self.get_visible_cols()
            all_columns = get_all_productgroup_columns()
            qs = self.get_queryset()
            if fmt == 'excel':
                return export_visible_columns_excel(qs, visible_columns, all_columns, self.get_export_value, 'Categorías', 'Categorias')
            return export_visible_columns_pdf(qs, visible_columns, all_columns, self.get_export_value, 'Listado de Categorías', 'Categorias')
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = ProductGroup.objects.annotate(product_count=Count('products'))
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(name__icontains=name)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        return qs

    def get_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('productgroup_visible_columns', PRODUCTGROUP_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_productgroup_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        if col_key == 'name':
            return obj.name
        elif col_key == 'product_count':
            return obj.product_count
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        return getattr(obj, col_key, '-')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.get_visible_cols()
        ctx['visible_columns'] = get_productgroup_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_productgroup_columns()
        ctx['visible_columns_list'] = visible_columns
        return ctx

class ProductGroupCreateView(PermissionRequiredAnyMixin, CreateView):
    permissions_required = ['billing.add_productgroup']
    model = ProductGroup; form_class = ProductGroupForm
    template_name = 'billing/product_group_form.html'
    success_url = reverse_lazy('billing:productgroup_list')

class ProductGroupUpdateView(PermissionRequiredAnyMixin, UpdateView):
    permissions_required = ['billing.change_productgroup']
    model = ProductGroup; form_class = ProductGroupForm
    template_name = 'billing/product_group_form.html'
    success_url = reverse_lazy('billing:productgroup_list')

class ProductGroupDetailView(PermissionRequiredAnyMixin, DetailView):
    permissions_required = ['billing.view_productgroup']
    model = ProductGroup
    template_name = 'billing/product_group_detail.html'
    context_object_name = 'group'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['products'] = self.object.products.all()[:10]
        ctx['product_count'] = self.object.products.count()
        return ctx

class ProductGroupDeleteView(PermissionRequiredAnyMixin, DeleteView):
    permissions_required = ['billing.delete_productgroup']
    model = ProductGroup
    template_name = 'billing/product_group_confirm_delete.html'
    success_url = reverse_lazy('billing:productgroup_list')

    def form_valid(self, form):
        from django.db.models.deletion import ProtectedError
        try:
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'No se puede eliminar esta categoría porque tiene productos asociados.')
            return redirect('billing:productgroup_list')

@login_required
def productgroup_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de categorías"""
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        is_valid, validated_columns = validate_productgroup_visible_columns(visible_columns)
        request.session['productgroup_visible_columns'] = validated_columns
        request.session.modified = True
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_productgroup_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === SUPPLIER (CBV) ===
class SupplierListView(PermissionRequiredAnyMixin, ExportListMixin, ListView):
    permissions_required = ['billing.view_supplier']
    model = Supplier
    template_name = 'billing/supplier_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Proveedores'
    export_fields = [
        ('Nombre', 'name'),
        ('Contacto', lambda o: o.contact_name or '-'),
        ('Email', lambda o: o.email or '-'),
        ('Teléfono', lambda o: o.phone or '-'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf') and not user_can_export(request, 'billing.export_supplier'):
            messages.error(request, 'No tienes permiso para exportar esta información.')
            fmt = None
        if fmt in ('excel', 'pdf'):
            visible_columns = self.get_visible_cols()
            all_columns = get_all_supplier_columns()
            qs = self.get_queryset()
            if fmt == 'excel':
                return export_visible_columns_excel(qs, visible_columns, all_columns, self.get_export_value, 'Proveedores', 'Proveedores')
            return export_visible_columns_pdf(qs, visible_columns, all_columns, self.get_export_value, 'Listado de Proveedores', 'Proveedores')
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Supplier.objects.annotate(product_count=Count('products'))
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(name__icontains=name)
        if email := g.get('email', '').strip():
            qs = qs.filter(email__icontains=email)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        return qs

    def get_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('supplier_visible_columns', SUPPLIER_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_supplier_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        if col_key == 'name':
            return obj.name
        elif col_key == 'contact_name':
            return obj.contact_name or '-'
        elif col_key == 'email':
            return obj.email or '-'
        elif col_key == 'phone':
            return obj.phone or '-'
        elif col_key == 'address':
            return obj.address or '-'
        elif col_key == 'product_count':
            return getattr(obj, 'product_count', obj.products.count())
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        return getattr(obj, col_key, '-')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.get_visible_cols()
        ctx['visible_columns'] = get_supplier_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_supplier_columns()
        ctx['visible_columns_list'] = visible_columns
        return ctx

class SupplierCreateView(PermissionRequiredAnyMixin, CreateView):
    permissions_required = ['billing.add_supplier']
    model = Supplier; form_class = SupplierForm
    template_name = 'billing/supplier_form.html'
    success_url = reverse_lazy('billing:supplier_list')

class SupplierUpdateView(PermissionRequiredAnyMixin, UpdateView):
    permissions_required = ['billing.change_supplier']
    model = Supplier; form_class = SupplierForm
    template_name = 'billing/supplier_form.html'
    success_url = reverse_lazy('billing:supplier_list')

class SupplierDetailView(PermissionRequiredAnyMixin, DetailView):
    permissions_required = ['billing.view_supplier']
    model = Supplier
    template_name = 'billing/supplier_detail.html'
    context_object_name = 'supplier'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['products'] = self.object.products.all()[:10]
        ctx['product_count'] = self.object.products.count()
        return ctx

class SupplierDeleteView(PermissionRequiredAnyMixin, DeleteView):
    permissions_required = ['billing.delete_supplier']
    model = Supplier
    template_name = 'billing/supplier_confirm_delete.html'
    success_url = reverse_lazy('billing:supplier_list')

    def form_valid(self, form):
        from django.db.models.deletion import ProtectedError
        try:
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'No se puede eliminar este proveedor porque tiene productos asociados.')
            return redirect('billing:supplier_list')

@login_required
def supplier_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de proveedores"""
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        is_valid, validated_columns = validate_supplier_visible_columns(visible_columns)
        request.session['supplier_visible_columns'] = validated_columns
        request.session.modified = True
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_supplier_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === PRODUCT (CBV) ===
class ProductListView(PermissionRequiredAnyMixin, ExportListMixin, ListView):
    permissions_required = ['billing.view_product']
    model = Product
    template_name = 'billing/product_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Productos'
    export_fields = [
        ('Nombre', 'name'),
        ('Marca', 'brand.name'),
        ('Grupo', 'group.name'),
        ('Precio', lambda o: f'{o.unit_price:.2f}'),
        ('Stock', 'stock'),
        ('Proveedores', lambda o: ', '.join(s.name for s in o.suppliers.all()) or '-'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf') and not user_can_export(request, 'billing.export_product'):
            messages.error(request, 'No tienes permiso para exportar esta información.')
            fmt = None
        if fmt == 'excel':
            return self.export_excel_with_visible_columns()
        if fmt == 'pdf':
            return self.export_pdf_with_visible_columns()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Product.objects.select_related('brand', 'group').prefetch_related('suppliers')
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(name__icontains=name)
        if brand := g.get('brand', ''):
            qs = qs.filter(brand_id=brand)
        if group := g.get('group', ''):
            qs = qs.filter(group_id=group)
        if price_min := g.get('price_min', '').strip():
            qs = qs.filter(unit_price__gte=price_min)
        if price_max := g.get('price_max', '').strip():
            qs = qs.filter(unit_price__lte=price_max)
        if stock_min := g.get('stock_min', '').strip():
            qs = qs.filter(stock__gte=stock_min)
        if stock_max := g.get('stock_max', '').strip():
            qs = qs.filter(stock__lte=stock_max)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        if supplier := g.get('supplier', ''):
            qs = qs.filter(suppliers__id=supplier).distinct()
        return qs

    def get_visible_columns(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('product_visible_columns', DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        """Obtener el valor de una columna específica"""
        col_config = get_all_columns()[col_key]
        field = col_config['field']
        
        if col_key == 'image':
            return 'Sí' if obj.image else 'No'
        elif col_key == 'brand':
            return obj.brand.name if obj.brand else '-'
        elif col_key == 'group':
            return obj.group.name if obj.group else '-'
        elif col_key == 'suppliers':
            return ', '.join(s.name for s in obj.suppliers.all()) or '-'
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'unit_price':
            return f'{obj.unit_price:.2f}'
        elif col_key == 'balance':
            return f'{obj.balance:.2f}'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        else:
            return getattr(obj, field, '-')

    def export_excel_with_visible_columns(self):
        """Exportar a Excel con columnas visibles"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        visible_columns = self.get_visible_columns()
        all_columns = get_all_columns()
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Productos'[:31]
        
        # Encabezados
        headers = [all_columns[col]['label'] for col in visible_columns]
        ws.append(headers)
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        
        # Datos
        rows = self.get_queryset()
        for row_obj in rows:
            row = [str(self.get_export_value(row_obj, col)) for col in visible_columns]
            ws.append(row)
        
        # Ajustar ancho de columnas
        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)
        
        # Respuesta
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Productos_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_pdf_with_visible_columns(self):
        """Exportar a PDF con columnas visibles"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape, portrait
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse
        
        visible_columns = self.get_visible_columns()
        all_columns = get_all_columns()
        
        response = HttpResponse(content_type='application/pdf')
        filename = f'Productos_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Determinar orientación basada en número de columnas
        pagesize = landscape(A4) if len(visible_columns) > 5 else A4
        
        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=0.8 * cm, rightMargin=0.8 * cm,
            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        )
        
        styles = getSampleStyleSheet()
        
        # Ajustar tamaño de fuente según número de columnas
        if len(visible_columns) > 8:
            font_size = 7
        elif len(visible_columns) > 5:
            font_size = 8
        else:
            font_size = 9
        
        cell_style = ParagraphStyle('cell', parent=styles['Normal'], 
                                   fontSize=font_size, leading=font_size + 2)
        head_style = ParagraphStyle('cellHead', parent=styles['Normal'],
                                   fontSize=font_size, leading=font_size + 2,
                                   textColor=colors.white, fontName='Helvetica-Bold')
        
        elements = [
            Paragraph('Listado de Productos', styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.3 * cm),
        ]
        
        # Encabezados
        headers = [all_columns[col]['label'] for col in visible_columns]
        data = [headers]
        
        # Datos
        for obj in self.get_queryset():
            row = [str(self.get_export_value(obj, col)) for col in visible_columns]
            data.append(row)
        
        # Crear tabla
        table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))
        
        # Estilos de tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['brands'] = Brand.objects.order_by('name')
        ctx['groups'] = ProductGroup.objects.order_by('name')
        ctx['suppliers'] = Supplier.objects.order_by('name')
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()
        
        # Obtener columnas visibles de la sesión o usar default
        visible_columns = self.request.session.get('product_visible_columns', DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns = validate_visible_columns(visible_columns)
        
        ctx['visible_columns'] = get_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_columns()
        ctx['visible_columns_list'] = visible_columns
        
        return ctx

class ProductCreateView(PermissionRequiredAnyMixin, CreateView):
    permissions_required = ['billing.add_product']
    model = Product; form_class = ProductForm
    template_name = 'billing/product_form.html'
    success_url = reverse_lazy('billing:product_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        for f in self.request.FILES.getlist('extra_images'):
            if f.size <= 5 * 1024 * 1024:
                ProductImage.objects.create(product=self.object, image=f)
        log_action(self.request, 'created', 'Product', self.object.pk, f'Producto creado: {self.object.name}')
        return response

class ProductUpdateView(PermissionRequiredAnyMixin, UpdateView):
    permissions_required = ['billing.change_product']
    model = Product; form_class = ProductForm
    template_name = 'billing/product_form.html'
    success_url = reverse_lazy('billing:product_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        for pk in self.request.POST.getlist('delete_image'):
            ProductImage.objects.filter(pk=pk, product=self.object).delete()
        for f in self.request.FILES.getlist('extra_images'):
            if f.size <= 5 * 1024 * 1024:
                ProductImage.objects.create(product=self.object, image=f)
        log_action(self.request, 'updated', 'Product', self.object.pk, f'Producto actualizado: {self.object.name}')
        return response

class ProductDeleteView(PermissionRequiredAnyMixin, DeleteView):
    permissions_required = ['billing.delete_product']
    model = Product
    template_name = 'billing/product_confirm_delete.html'
    success_url = reverse_lazy('billing:product_list')

    def form_valid(self, form):
        pk, name = self.object.pk, str(self.object)
        response = super().form_valid(form)
        log_action(self.request, 'deleted', 'Product', pk, f'Producto eliminado: {name}')
        return response


class ProductDetailView(PermissionRequiredAnyMixin, DetailView):
    permissions_required = ['billing.view_product']
    model = Product
    template_name = 'billing/product_detail.html'
    context_object_name = 'product'


@permission_required_any('billing.change_product')
def product_update_image(request, pk):
    """Actualizar imagen del producto via AJAX"""
    from django.http import JsonResponse
    from django.views.decorators.http import require_http_methods
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    product = get_object_or_404(Product, pk=pk)
    
    if 'image' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No se envió imagen'}, status=400)
    
    image_file = request.FILES['image']
    
    # Validar tipo de archivo
    valid_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
    if image_file.content_type not in valid_types:
        return JsonResponse({'success': False, 'error': 'Tipo de archivo no válido'}, status=400)
    
    # Validar tamaño (5MB)
    if image_file.size > 5 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Archivo muy grande (máx: 5MB)'}, status=400)
    
    try:
        product.image = image_file
        product.save()
        return JsonResponse({
            'success': True,
            'image_url': product.image.url,
            'message': 'Imagen actualizada correctamente'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def product_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de productos"""
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        
        # Validar columnas
        is_valid, validated_columns = validate_visible_columns(visible_columns)
        
        # Guardar en sesión
        request.session['product_visible_columns'] = validated_columns
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === CUSTOMER (CBV) ===
class CustomerListView(PermissionRequiredAnyMixin, ExportListMixin, ListView):
    permissions_required = ['billing.view_customer']
    model = Customer
    template_name = 'billing/customer_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Clientes'
    export_fields = [
        ('DNI/RUC', 'dni'),
        ('Apellido', 'last_name'),
        ('Nombre', 'first_name'),
        ('Email', lambda o: o.email or '-'),
        ('Teléfono', lambda o: o.phone or '-'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf') and not user_can_export(request, 'billing.export_customer'):
            messages.error(request, 'No tienes permiso para exportar esta información.')
            fmt = None
        if fmt == 'excel':
            return self.export_excel_with_visible_columns()
        if fmt == 'pdf':
            return self.export_pdf_with_visible_columns()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Customer.objects.select_related('profile')
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(first_name__icontains=name) | qs.filter(last_name__icontains=name)
        if dni := g.get('dni', '').strip():
            qs = qs.filter(dni__icontains=dni)
        if email := g.get('email', '').strip():
            qs = qs.filter(email__icontains=email)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        return qs

    def get_customer_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('customer_visible_columns', CUSTOMER_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_customer_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        """Obtener el valor de una columna específica"""
        if col_key == 'dni':
            return obj.dni
        elif col_key == 'last_name':
            return obj.last_name
        elif col_key == 'first_name':
            return obj.first_name
        elif col_key == 'email':
            return obj.email or '-'
        elif col_key == 'phone':
            return obj.phone or '-'
        elif col_key == 'address':
            return obj.address or '-'
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'taxpayer_type':
            try:
                return obj.profile.get_taxpayer_type_display()
            except Exception:
                return '-'
        elif col_key == 'payment_terms':
            try:
                return obj.profile.get_payment_terms_display()
            except Exception:
                return '-'
        elif col_key == 'credit_limit':
            try:
                return f'{obj.profile.credit_limit:.2f}'
            except Exception:
                return '-'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        else:
            return getattr(obj, col_key, '-')

    def export_excel_with_visible_columns(self):
        """Exportar a Excel con columnas visibles"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        visible_columns = self.get_customer_visible_cols()
        all_columns = get_all_customer_columns()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'[:31]

        headers = [all_columns[col]['label'] for col in visible_columns]
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for row_obj in self.get_queryset():
            row = [str(self.get_export_value(row_obj, col)) for col in visible_columns]
            ws.append(row)

        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Clientes_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_pdf_with_visible_columns(self):
        """Exportar a PDF con columnas visibles"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse

        visible_columns = self.get_customer_visible_cols()
        all_columns = get_all_customer_columns()

        response = HttpResponse(content_type='application/pdf')
        filename = f'Clientes_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        pagesize = landscape(A4) if len(visible_columns) > 5 else A4

        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=0.8 * cm, rightMargin=0.8 * cm,
            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        )

        styles = getSampleStyleSheet()

        if len(visible_columns) > 8:
            font_size = 7
        elif len(visible_columns) > 5:
            font_size = 8
        else:
            font_size = 9

        cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                   fontSize=font_size, leading=font_size + 2)
        head_style = ParagraphStyle('cellHead', parent=styles['Normal'],
                                   fontSize=font_size, leading=font_size + 2,
                                   textColor=colors.white, fontName='Helvetica-Bold')

        elements = [
            Paragraph('Listado de Clientes', styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.3 * cm),
        ]

        headers = [all_columns[col]['label'] for col in visible_columns]
        data = [headers]

        for obj in self.get_queryset():
            row = [str(self.get_export_value(obj, col)) for col in visible_columns]
            data.append(row)

        table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.request.session.get('customer_visible_columns', CUSTOMER_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns = validate_customer_visible_columns(visible_columns)

        ctx['visible_columns'] = get_customer_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_customer_columns()
        ctx['visible_columns_list'] = visible_columns

        return ctx

class CustomerCreateView(PermissionRequiredAnyMixin, CreateView):
    permissions_required = ['billing.add_customer']
    model = Customer; form_class = CustomerForm
    template_name = 'billing/customer_form.html'
    success_url = reverse_lazy('billing:customer_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request, 'created', 'Customer', self.object.pk, f'Cliente creado: {self.object.full_name}')
        return response

class CustomerUpdateView(PermissionRequiredAnyMixin, UpdateView):
    permissions_required = ['billing.change_customer']
    model = Customer; form_class = CustomerForm
    template_name = 'billing/customer_form.html'
    success_url = reverse_lazy('billing:customer_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request, 'updated', 'Customer', self.object.pk, f'Cliente actualizado: {self.object.full_name}')
        return response

class CustomerDetailView(PermissionRequiredAnyMixin, DetailView):
    permissions_required = ['billing.view_customer']
    model = Customer
    template_name = 'billing/customer_detail.html'
    context_object_name = 'customer'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        customer = self.object
        ctx['invoices'] = customer.invoices.order_by('-invoice_date')[:10]
        try:
            ctx['profile'] = customer.profile
        except CustomerProfile.DoesNotExist:
            ctx['profile'] = None
        return ctx

class CustomerDeleteView(PermissionRequiredAnyMixin, DeleteView):
    permissions_required = ['billing.delete_customer']
    model = Customer
    template_name = 'billing/customer_confirm_delete.html'
    success_url = reverse_lazy('billing:customer_list')

    def form_valid(self, form):
        pk, name = self.object.pk, self.object.full_name
        response = super().form_valid(form)
        log_action(self.request, 'deleted', 'Customer', pk, f'Cliente eliminado: {name}')
        return response


@login_required
def customer_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de clientes"""
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])

        is_valid, validated_columns = validate_customer_visible_columns(visible_columns)

        request.session['customer_visible_columns'] = validated_columns
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_customer_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === INVOICE (CBV) ===
class InvoiceListView(PermissionRequiredAnyMixin, ExportListMixin, ListView):
    permissions_required = ['billing.view_invoice']
    model = Invoice
    template_name = 'billing/invoice_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Facturas'
    export_fields = [
        ('#', 'id'),
        ('Cliente', lambda o: str(o.customer)),
        ('Fecha', lambda o: o.invoice_date.strftime('%d/%m/%Y %H:%M')),
        ('Subtotal', lambda o: f'{o.subtotal:.2f}'),
        ('Impuesto', lambda o: f'{o.tax:.2f}'),
        ('Total', lambda o: f'{o.total:.2f}'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf') and not user_can_export(request, 'billing.export_invoice'):
            messages.error(request, 'No tienes permiso para exportar esta información.')
            fmt = None
        if fmt == 'excel':
            return self.export_excel_with_visible_columns()
        if fmt == 'pdf':
            return self.export_pdf_with_visible_columns()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Invoice.objects.select_related('customer')
        g = self.request.GET
        if customer := g.get('customer', '').strip():
            qs = qs.filter(
                customer__first_name__icontains=customer
            ) | qs.filter(customer__last_name__icontains=customer)
        if date_from := g.get('date_from', '').strip():
            qs = qs.filter(invoice_date__date__gte=date_from)
        if date_to := g.get('date_to', '').strip():
            qs = qs.filter(invoice_date__date__lte=date_to)
        if total_min := g.get('total_min', '').strip():
            qs = qs.filter(total__gte=total_min)
        if total_max := g.get('total_max', '').strip():
            qs = qs.filter(total__lte=total_max)
        if estado := g.get('estado', '').strip():
            if estado in ('pendiente', 'parcial', 'pagada', 'anulada'):
                qs = qs.filter(estado=estado)
        return qs

    def get_invoice_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('invoice_visible_columns', INVOICE_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_invoice_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        """Obtener el valor de una columna específica"""
        if col_key == 'id':
            return obj.id
        elif col_key == 'customer':
            return str(obj.customer)
        elif col_key == 'customer_dni':
            return obj.customer.dni if obj.customer else '-'
        elif col_key == 'invoice_date':
            return obj.invoice_date.strftime('%d/%m/%Y %H:%M') if obj.invoice_date else '-'
        elif col_key == 'num_items':
            return obj.details.count()
        elif col_key == 'subtotal':
            return f'{obj.subtotal:.2f}'
        elif col_key == 'tax':
            return f'{obj.tax:.2f}'
        elif col_key == 'total':
            return f'{obj.total:.2f}'
        elif col_key == 'is_active':
            return 'Activa' if obj.is_active else 'Anulada'
        elif col_key == 'estado':
            return obj.get_estado_display()
        elif col_key == 'saldo':
            return f'{obj.saldo:.2f}'
        elif col_key == 'tipo_pago':
            return obj.get_tipo_pago_display()
        else:
            return getattr(obj, col_key, '-')

    def export_excel_with_visible_columns(self):
        """Exportar a Excel con columnas visibles"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        visible_columns = self.get_invoice_visible_cols()
        all_columns = get_all_invoice_columns()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Facturas'[:31]

        headers = [all_columns[col]['label'] for col in visible_columns]
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for row_obj in self.get_queryset():
            row = [str(self.get_export_value(row_obj, col)) for col in visible_columns]
            ws.append(row)

        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Facturas_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_pdf_with_visible_columns(self):
        """Exportar a PDF con columnas visibles"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse

        visible_columns = self.get_invoice_visible_cols()
        all_columns = get_all_invoice_columns()

        response = HttpResponse(content_type='application/pdf')
        filename = f'Facturas_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        pagesize = landscape(A4) if len(visible_columns) > 5 else A4

        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=0.8 * cm, rightMargin=0.8 * cm,
            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        )

        styles = getSampleStyleSheet()

        if len(visible_columns) > 8:
            font_size = 7
        elif len(visible_columns) > 5:
            font_size = 8
        else:
            font_size = 9

        elements = [
            Paragraph('Listado de Facturas', styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.3 * cm),
        ]

        headers = [all_columns[col]['label'] for col in visible_columns]
        data = [headers]

        for obj in self.get_queryset():
            row = [str(self.get_export_value(obj, col)) for col in visible_columns]
            data.append(row)

        table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.request.session.get('invoice_visible_columns', INVOICE_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns = validate_invoice_visible_columns(visible_columns)

        ctx['visible_columns'] = get_invoice_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_invoice_columns()
        ctx['visible_columns_list'] = visible_columns

        return ctx

@permission_required_any('billing.add_invoice')
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        formset = InvoiceDetailFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            # Validar stock antes de guardar
            stock_errors = []
            for detail_form in formset:
                if detail_form.cleaned_data and not detail_form.cleaned_data.get('DELETE', False):
                    product = detail_form.cleaned_data.get('product')
                    quantity = detail_form.cleaned_data.get('quantity', 0)
                    if product and quantity:
                        if quantity > product.stock:
                            stock_errors.append(
                                f'"{product.name}" solo tiene {product.stock} unidades en stock, '
                                f'pero intentas vender {quantity}.'
                            )
            
            if stock_errors:
                for error in stock_errors:
                    messages.error(request, error)
            else:
                try:
                    with transaction.atomic():
                        invoice = form.save()
                        formset.instance = invoice
                        formset.save()

                        # Descuento atómico de stock: si entre la validación de
                        # arriba y este punto otra factura ya consumió el stock
                        # (dos ventas casi simultáneas del mismo producto), esta
                        # actualización afecta 0 filas y abortamos toda la
                        # transacción en vez de dejar stock en negativo.
                        for detail in invoice.details.all():
                            updated = Product.objects.filter(
                                pk=detail.product_id,
                                stock__gte=detail.quantity
                            ).update(stock=F('stock') - detail.quantity)
                            if not updated:
                                raise ValueError(
                                    f'"{detail.product.name}" ya no tiene stock '
                                    f'suficiente para vender {detail.quantity} unidades. '
                                    f'Otra venta lo consumió justo antes.'
                                )

                        subtotal = sum(d.subtotal for d in invoice.details.all())
                        invoice.subtotal = subtotal
                        invoice.tax = subtotal * Decimal('0.15')
                        invoice.total = invoice.subtotal + invoice.tax

                        if invoice.tipo_pago == 'credito':
                            from billing.services import check_credit_limit
                            check_credit_limit(invoice.customer, invoice.total)
                            invoice.saldo = invoice.total
                            invoice.estado = 'pendiente'

                        invoice.save()

                        # Si es a crédito y se indicó número de cuotas, se genera
                        # el cronograma de una vez (si no, queda pendiente y se
                        # puede generar después desde el detalle de la factura).
                        numero_cuotas = form.cleaned_data.get('numero_cuotas')
                        if invoice.tipo_pago == 'credito' and numero_cuotas:
                            from creditos_ventas.services import generar_cuotas
                            generar_cuotas(invoice, numero_cuotas)
                except ValueError as e:
                    messages.error(request, str(e))
                else:
                    log_action(request, 'created', 'Invoice', invoice.id, f'Factura #{invoice.id} creada. Total: ${invoice.total}')
                    mensaje = f'Factura #{invoice.id} creada! Total: ${invoice.total}'
                    if invoice.tipo_pago == 'credito' and form.cleaned_data.get('numero_cuotas'):
                        mensaje += f'. Se generaron {form.cleaned_data["numero_cuotas"]} cuotas.'
                    messages.success(request, mensaje)
                    return redirect('billing:invoice_detail', pk=invoice.id)
    else:
        form = InvoiceForm()
        formset = InvoiceDetailFormSet()
    return render(request, 'billing/invoice_form.html', {
        'form': form, 'formset': formset, 'title': 'Nueva Factura',
        'today': timezone.localtime().strftime('%d/%m/%Y'),
    })


@permission_required_any('billing.view_product')
def api_product_info(request, pk):
    """API para obtener precio y stock de un producto (usado en el formulario de factura)."""
    from django.http import JsonResponse
    product = get_object_or_404(Product, pk=pk)
    return JsonResponse({
        'id': product.id,
        'name': product.name,
        'unit_price': str(product.unit_price),
        'stock': product.stock,
    })


class InvoiceDetailView(PermissionRequiredAnyMixin, DetailView):
    permissions_required = ['billing.view_invoice']
    model = Invoice
    template_name = 'billing/invoice_detail.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        invoice = self.object
        details = invoice.details.select_related('product', 'product__brand').all()
        ctx['details'] = details
        ctx['total_units'] = sum(d.quantity for d in details)
        ctx['payments'] = invoice.payments.select_related('registered_by').all()
        ctx['tiene_cuotas'] = invoice.cuotas.exists()
        ctx['tiene_pagos_libres'] = invoice.payments.exists() or invoice.cobros.exists()
        return ctx


@login_required
def invoice_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de facturas"""
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])

        is_valid, validated_columns = validate_invoice_visible_columns(visible_columns)

        request.session['invoice_visible_columns'] = validated_columns
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_invoice_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@permission_required_any('billing.delete_invoice')
def invoice_void(request, pk):
    """Anula una factura (soft-delete): estado='anulada', is_active=False, restaura stock.
    No elimina el registro — mantiene el historial y la auditoría intactos."""
    invoice = get_object_or_404(Invoice, pk=pk)

    if invoice.estado == 'anulada':
        messages.warning(request, 'Esta factura ya está anulada.')
        return redirect('billing:invoice_detail', pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            for detail in invoice.details.select_related('product').all():
                Product.objects.filter(pk=detail.product_id).update(
                    stock=F('stock') + detail.quantity
                )
            invoice.estado = 'anulada'
            invoice.is_active = False
            invoice.saldo = 0
            invoice.save(update_fields=['estado', 'is_active', 'saldo'])
        log_action(request, 'deleted', 'Invoice', pk, f'Factura #{pk} anulada (soft-delete) y stock restaurado')
        messages.success(request, f'Factura #{pk} anulada correctamente. El stock fue restaurado.')
        return redirect('billing:invoice_list')

    return render(request, 'billing/invoice_confirm_void.html', {'invoice': invoice})


class InvoiceDeleteView(PermissionRequiredAnyMixin, DeleteView):
    permissions_required = ['billing.delete_invoice']
    model = Invoice
    template_name = 'billing/invoice_confirm_delete.html'
    success_url = reverse_lazy('billing:invoice_list')

    def form_valid(self, form):
        from django.db.models.deletion import ProtectedError
        pk = self.object.pk
        try:
            with transaction.atomic():
                # Restaurar stock de cada línea antes de eliminar
                for detail in self.object.details.select_related('product').all():
                    Product.objects.filter(pk=detail.product_id).update(
                        stock=F('stock') + detail.quantity
                    )
                response = super().form_valid(form)
        except ProtectedError:
            messages.error(
                self.request,
                'No se puede eliminar esta factura porque tiene pagos, cobros o cuotas registradas. '
                'Elimínalos primero.'
            )
            return redirect('billing:invoice_detail', pk=pk)
        log_action(self.request, 'deleted', 'Invoice', pk, f'Factura #{pk} eliminada y stock restaurado')
        return response


@permission_required_any('billing.export_invoice')
def invoice_pdf(request, pk):
    """Server-generated PDF for a single invoice (opens inline in browser)."""
    from billing.services import build_invoice_pdf
    from django.http import HttpResponse

    invoice = get_object_or_404(
        Invoice.objects.select_related('customer').prefetch_related('details__product__brand', 'payments__registered_by'),
        pk=pk,
    )
    buffer = build_invoice_pdf(invoice)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="factura_{invoice.id:05d}.pdf"'
    return response


@permission_required_any('billing.change_invoice')
def register_payment(request, pk):
    """Registrar un pago (total o parcial) en una factura de crédito."""
    from billing.services import register_invoice_payment
    invoice = get_object_or_404(Invoice, pk=pk, is_active=True)

    if invoice.tipo_pago != 'credito' or invoice.saldo <= 0:
        messages.warning(request, 'Esta factura no tiene saldo pendiente.')
        return redirect('billing:invoice_detail', pk=pk)

    if invoice.cuotas.exists():
        messages.error(
            request,
            'Esta factura tiene un cronograma de cuotas generado. '
            'Registra los pagos desde el módulo de cuotas.'
        )
        return redirect('creditos_ventas:cuota_list', factura_id=invoice.id)

    if request.method == 'POST':
        form = InvoicePaymentForm(request.POST)
        if form.is_valid():
            try:
                payment = register_invoice_payment(
                    invoice=invoice,
                    amount=form.cleaned_data['amount'],
                    method=form.cleaned_data['method'],
                    user=request.user,
                    notes=form.cleaned_data.get('notes', ''),
                )
                log_action(
                    request, 'updated', 'Invoice', invoice.pk,
                    f'Pago de ${payment.amount:.2f} registrado en Factura #{invoice.pk}. '
                    f'Saldo restante: ${invoice.saldo:.2f}',
                )
                messages.success(
                    request,
                    f'Pago de ${payment.amount:.2f} registrado. Saldo restante: ${invoice.saldo:.2f}',
                )
            except ValueError as e:
                messages.error(request, str(e))
                return render(request, 'billing/invoice_payment_form.html', {
                    'form': form,
                    'invoice': invoice,
                })
            return redirect('billing:invoice_detail', pk=pk)
    else:
        form = InvoicePaymentForm(initial={'amount': invoice.saldo})

    return render(request, 'billing/invoice_payment_form.html', {
        'form': form,
        'invoice': invoice,
    })


# ─────────────────────────────────────────────
# Importar productos desde Excel
# ─────────────────────────────────────────────

@permission_required_any('billing.add_product')
def product_import(request):
    """Vista principal de importación de productos desde .xlsx.
    Flujo: el usuario sube el archivo → validamos fila por fila →
    mostramos una previsualización con errores marcados → el usuario
    confirma y guardamos solo las filas válidas."""

    if request.method != 'POST' or 'excel_file' not in request.FILES:
        return render(request, 'billing/product_import.html')

    excel_file = request.FILES['excel_file']
    if not excel_file.name.endswith('.xlsx'):
        messages.error(request, 'El archivo debe ser .xlsx')
        return render(request, 'billing/product_import.html')

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, 'No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx).')
        return render(request, 'billing/product_import.html')

    EXPECTED_HEADERS = ['nombre', 'marca', 'categoria', 'precio', 'stock', 'descripcion']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, 'El archivo está vacío.')
        return render(request, 'billing/product_import.html')

    # Normalizar headers (minúsculas, sin espacios)
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    missing = [h for h in EXPECTED_HEADERS[:5] if h not in headers]  # descripcion es opcional
    if missing:
        messages.error(request, f'Faltan columnas requeridas: {", ".join(missing)}. Descarga la plantilla para ver el formato correcto.')
        return render(request, 'billing/product_import.html')

    def col(row_dict, name):
        return row_dict.get(name, '')

    preview_rows = []
    for i, raw_row in enumerate(rows[1:], start=2):
        row_dict = {headers[j]: raw_row[j] for j in range(len(headers)) if j < len(raw_row)}
        errors = []

        nombre = str(col(row_dict, 'nombre') or '').strip()
        marca_name = str(col(row_dict, 'marca') or '').strip()
        cat_name = str(col(row_dict, 'categoria') or '').strip()
        precio_raw = col(row_dict, 'precio')
        stock_raw = col(row_dict, 'stock')
        descripcion = str(col(row_dict, 'descripcion') or '').strip()

        # Ignorar filas completamente vacías
        if not any([nombre, marca_name, cat_name, precio_raw, stock_raw]):
            continue

        if not nombre:
            errors.append('Nombre requerido')
        if not marca_name:
            errors.append('Marca requerida')
        if not cat_name:
            errors.append('Categoría requerida')

        try:
            precio = Decimal(str(precio_raw).replace(',', '.')).quantize(Decimal('0.01'))
            if precio < 0:
                errors.append('Precio no puede ser negativo')
        except Exception:
            precio = None
            errors.append('Precio inválido (usa número, ej: 12.50)')

        try:
            stock = int(float(str(stock_raw)))
            if stock < 0:
                errors.append('Stock no puede ser negativo')
        except (ValueError, TypeError):
            stock = None
            errors.append('Stock inválido (usa número entero, ej: 10)')

        preview_rows.append({
            'fila': i,
            'nombre': nombre,
            'marca': marca_name,
            'categoria': cat_name,
            'precio': precio,
            'stock': stock,
            'descripcion': descripcion,
            'errores': errors,
            'valido': len(errors) == 0,
        })

    if not preview_rows:
        messages.warning(request, 'El archivo no tiene filas de datos (solo encabezados).')
        return render(request, 'billing/product_import.html')

    # Si el usuario ya confirmó la importación
    if 'confirmar' in request.POST:
        imported = 0
        skipped = 0
        with transaction.atomic():
            for row in preview_rows:
                if not row['valido']:
                    skipped += 1
                    continue
                brand = Brand.objects.filter(name__iexact=row['marca']).first()
                if not brand:
                    brand, _ = Brand.objects.get_or_create(name=row['marca'], defaults={'is_active': True})
                group = ProductGroup.objects.filter(name__iexact=row['categoria']).first()
                if not group:
                    group, _ = ProductGroup.objects.get_or_create(name=row['categoria'], defaults={'is_active': True})
                Product.objects.update_or_create(
                    name=row['nombre'],
                    brand=brand,
                    defaults={
                        'group': group,
                        'unit_price': row['precio'],
                        'stock': row['stock'],
                        'description': row['descripcion'],
                        'is_active': True,
                    }
                )
                imported += 1

        messages.success(request, f'Importación completada: {imported} producto(s) importado(s), {skipped} fila(s) con errores omitida(s).')
        return redirect('billing:product_list')

    validas = sum(1 for r in preview_rows if r['valido'])
    invalidas = len(preview_rows) - validas
    return render(request, 'billing/product_import.html', {
        'preview_rows': preview_rows,
        'validas': validas,
        'invalidas': invalidas,
    })


@permission_required_any('billing.view_product')
def product_import_template(request):
    """Descarga la plantilla Excel vacía con el formato correcto."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_TEXT
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # Encabezados
    headers = ['nombre', 'marca', 'categoria', 'precio', 'stock', 'descripcion']
    header_fill = PatternFill('solid', start_color='231A10')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    border = Border(
        bottom=Side(style='thin', color='B5441B'),
        right=Side(style='thin', color='DDD3C5'),
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Filas de ejemplo
    examples = [
        ['Camiseta básica', 'Nike', 'Ropa', 29.99, 50, 'Algodón 100%, talla M'],
        ['Laptop Pro 15"', 'LG', 'Electrónica', 899.00, 10, ''],
        ['Silla de oficina', 'IKEA', 'Hogar', 149.50, 5, 'Con ruedas, regulable'],
    ]
    example_font = Font(color='555555', italic=True, name='Arial', size=10)
    example_fill = PatternFill('solid', start_color='F8F3EE')
    for row_num, row_data in enumerate(examples, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.alignment = Alignment(vertical='center')

    # Nota explicativa
    note_row = len(examples) + 3
    ws.cell(row=note_row, column=1, value='INSTRUCCIONES:').font = Font(bold=True, name='Arial', size=10)
    notes = [
        '• Las columnas nombre, marca, categoria, precio y stock son obligatorias.',
        '• Si la marca o categoría no existe, se crea automáticamente.',
        '• Si el producto ya existe (mismo nombre), se actualiza.',
        '• Borra las filas de ejemplo antes de importar.',
        '• No cambies los nombres de los encabezados de la fila 1.',
    ]
    for i, note in enumerate(notes, note_row + 1):
        ws.cell(row=i, column=1, value=note).font = Font(color='8B7355', name='Arial', size=9)

    ws.merge_cells(f'A{note_row}:F{note_row}')
    for i in range(note_row + 1, note_row + 1 + len(notes)):
        ws.merge_cells(f'A{i}:F{i}')

    # Anchos de columna
    widths = [30, 20, 20, 12, 10, 35]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    ws.row_dimensions[1].height = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
# Importar clientes desde Excel
# ─────────────────────────────────────────────

@permission_required_any('billing.add_customer')
def customer_import(request):
    """Importa clientes desde un .xlsx con columnas:
    cedula, nombre, apellido, email, telefono, direccion.
    Si el cliente ya existe (misma cédula), actualiza sus datos."""

    if request.method != 'POST' or 'excel_file' not in request.FILES:
        return render(request, 'billing/customer_import.html')

    excel_file = request.FILES['excel_file']
    if not excel_file.name.endswith('.xlsx'):
        messages.error(request, 'El archivo debe ser .xlsx')
        return render(request, 'billing/customer_import.html')

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, 'No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx).')
        return render(request, 'billing/customer_import.html')

    REQUIRED = ['cedula', 'nombre', 'apellido']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, 'El archivo está vacío.')
        return render(request, 'billing/customer_import.html')

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    missing = [h for h in REQUIRED if h not in headers]
    if missing:
        messages.error(request, f'Faltan columnas requeridas: {", ".join(missing)}. Descarga la plantilla para ver el formato correcto.')
        return render(request, 'billing/customer_import.html')

    from shared.validators import validate_cedula_ec
    from django.core.exceptions import ValidationError as DjangoValidationError
    import re

    preview_rows = []
    for i, raw_row in enumerate(rows[1:], start=2):
        row_dict = {headers[j]: raw_row[j] for j in range(len(headers)) if j < len(raw_row)}
        errors = []

        cedula    = str(row_dict.get('cedula', '') or '').strip()
        nombre    = str(row_dict.get('nombre', '') or '').strip()
        apellido  = str(row_dict.get('apellido', '') or '').strip()
        email     = str(row_dict.get('email', '') or '').strip()
        telefono  = str(row_dict.get('telefono', '') or '').strip()
        direccion = str(row_dict.get('direccion', '') or '').strip()

        if not any([cedula, nombre, apellido]):
            continue

        if not cedula:
            errors.append('Cédula requerida')
        else:
            try:
                validate_cedula_ec(cedula)
            except DjangoValidationError as e:
                errors.append(str(e.message))

        if not nombre:
            errors.append('Nombre requerido')
        if not apellido:
            errors.append('Apellido requerido')

        if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            errors.append('Email inválido')

        exists = Customer.objects.filter(dni=cedula).exists() if cedula and not errors else False

        preview_rows.append({
            'fila': i,
            'cedula': cedula,
            'nombre': nombre,
            'apellido': apellido,
            'email': email,
            'telefono': telefono,
            'direccion': direccion,
            'errores': errors,
            'valido': len(errors) == 0,
            'accion': 'Actualizar' if exists else 'Crear',
        })

    if not preview_rows:
        messages.warning(request, 'El archivo no tiene filas de datos.')
        return render(request, 'billing/customer_import.html')

    if 'confirmar' in request.POST:
        imported = updated = skipped = 0
        with transaction.atomic():
            for row in preview_rows:
                if not row['valido']:
                    skipped += 1
                    continue
                obj, created = Customer.objects.update_or_create(
                    dni=row['cedula'],
                    defaults={
                        'first_name': row['nombre'],
                        'last_name':  row['apellido'],
                        'email':      row['email'] or None,
                        'phone':      row['telefono'] or None,
                        'address':    row['direccion'] or None,
                        'is_active':  True,
                    }
                )
                if created:
                    imported += 1
                else:
                    updated += 1

        messages.success(
            request,
            f'Importación completada: {imported} cliente(s) nuevo(s), '
            f'{updated} actualizado(s), {skipped} fila(s) con errores omitida(s).'
        )
        return redirect('billing:customer_list')

    validas   = sum(1 for r in preview_rows if r['valido'])
    invalidas = len(preview_rows) - validas
    return render(request, 'billing/customer_import.html', {
        'preview_rows': preview_rows,
        'validas':   validas,
        'invalidas': invalidas,
    })


@permission_required_any('billing.view_customer')
def customer_import_template(request):
    """Descarga la plantilla Excel para importar clientes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_TEXT
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    headers = ['cedula', 'nombre', 'apellido', 'email', 'telefono', 'direccion']
    header_fill = PatternFill('solid', start_color='231A10')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    border = Border(
        bottom=Side(style='thin', color='B5441B'),
        right=Side(style='thin', color='DDD3C5'),
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    examples = [
        ['1710034065', 'Juan', 'Pérez', 'juan@gmail.com', '0999999999', 'Quito, Pichincha'],
        ['1720034066', 'María', 'López', 'maria@gmail.com', '0988888888', 'Guayaquil, Guayas'],
        ['1730034067', 'Carlos', 'Gómez', '', '0977777777', ''],
    ]
    example_font = Font(color='555555', italic=True, name='Arial', size=10)
    example_fill = PatternFill('solid', start_color='F8F3EE')
    for row_num, row_data in enumerate(examples, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.alignment = Alignment(vertical='center')

    note_row = len(examples) + 3
    ws.cell(row=note_row, column=1, value='INSTRUCCIONES:').font = Font(bold=True, name='Arial', size=10)
    notes = [
        '• Las columnas cedula, nombre y apellido son obligatorias.',
        '• Email, teléfono y dirección son opcionales.',
        '• Si el cliente ya existe (misma cédula), se actualizan sus datos.',
        '• La cédula debe ser válida (10 dígitos) o RUC (13 dígitos).',
        '• Borra las filas de ejemplo antes de importar.',
        '• No cambies los nombres de los encabezados de la fila 1.',
    ]
    for i, note in enumerate(notes, note_row + 1):
        ws.cell(row=i, column=1, value=note).font = Font(color='8B7355', name='Arial', size=9)

    ws.merge_cells(f'A{note_row}:F{note_row}')
    for i in range(note_row + 1, note_row + 1 + len(notes)):
        ws.merge_cells(f'A{i}:F{i}')

    widths = [15, 20, 20, 30, 15, 35]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    ws.row_dimensions[1].height = 22

    # Formatear cédula y teléfono como texto para evitar notación científica y pérdida del 0
    from openpyxl.utils import get_column_letter
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = FORMAT_TEXT
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = FORMAT_TEXT

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_clientes.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
# Reportes
# ─────────────────────────────────────────────

def _get_report_dates(request):
    """Extrae fecha_desde y fecha_hasta del request (GET o POST).
    Por defecto: el mes actual."""
    from datetime import date
    today = date.today()
    default_from = today.replace(day=1).isoformat()
    default_to   = today.isoformat()
    date_from_str = request.GET.get('date_from', default_from)
    date_to_str   = request.GET.get('date_to',   default_to)
    try:
        from datetime import datetime
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        from datetime import date
        date_from = date.today().replace(day=1)
        date_to   = date.today()
    return date_from, date_to


def _sales_queryset(date_from, date_to):
    return (
        Invoice.objects
        .filter(
            invoice_date__date__gte=date_from,
            invoice_date__date__lte=date_to,
            is_active=True,
        )
        .exclude(estado='anulada')
        .select_related('customer')
        .prefetch_related('details__product')
        .order_by('-invoice_date')
    )


def _sales_summary(invoices):
    from decimal import Decimal
    subtotal = sum(inv.subtotal for inv in invoices)
    tax      = sum(inv.tax for inv in invoices)
    total    = sum(inv.total for inv in invoices)
    count    = len(invoices)

    # Top 5 productos más vendidos
    product_totals = {}
    for inv in invoices:
        for d in inv.details.all():
            key = d.product.name
            if key not in product_totals:
                product_totals[key] = {'qty': 0, 'total': Decimal('0')}
            product_totals[key]['qty']   += d.quantity
            product_totals[key]['total'] += d.subtotal

    top_products = sorted(
        product_totals.items(), key=lambda x: x[1]['total'], reverse=True
    )[:5]

    return {
        'count': count,
        'subtotal': subtotal,
        'tax': tax,
        'total': total,
        'top_products': top_products,
    }


@permission_required_any('billing.view_invoice', 'billing.descargar_reportes_financieros')
def report_sales(request):
    date_from, date_to = _get_report_dates(request)
    invoices = list(_sales_queryset(date_from, date_to))
    summary  = _sales_summary(invoices)
    return render(request, 'billing/report_sales.html', {
        'invoices':  invoices,
        'summary':   summary,
        'date_from': date_from,
        'date_to':   date_to,
    })


@permission_required_any('billing.descargar_reportes_financieros')
def report_sales_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_TEXT
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    date_from, date_to = _get_report_dates(request)
    invoices = list(_sales_queryset(date_from, date_to))
    summary  = _sales_summary(invoices)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'Reporte de Ventas — {date_from.strftime("%d/%m/%Y")} al {date_to.strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13, name='Arial', color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='231A10')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Resumen
    ws['A3'] = 'Resumen'
    ws['A3'].font = Font(bold=True, name='Arial', size=11)
    summary_data = [
        ('Total facturas', summary['count']),
        ('Subtotal', f"${summary['subtotal']}"),
        ('IVA (15%)', f"${summary['tax']}"),
        ('Total general', f"${summary['total']}"),
    ]
    for i, (label, value) in enumerate(summary_data, 4):
        ws.cell(row=i, column=1, value=label).font = Font(name='Arial', size=10, bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = Font(name='Arial', size=10, color='B5441B')

    # Encabezados detalle
    headers = ['#', 'Fecha', 'Cliente', 'Cédula', 'Subtotal', 'IVA', 'Total']
    row_start = 10
    header_fill = PatternFill('solid', start_color='231A10')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_start, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Filas
    alt_fill = PatternFill('solid', start_color='F8F3EE')
    for i, inv in enumerate(invoices, row_start + 1):
        fill = alt_fill if i % 2 == 0 else None
        data = [
            inv.id,
            inv.invoice_date.strftime('%d/%m/%Y %H:%M'),
            inv.customer.full_name,
            inv.customer.dni,
            float(inv.subtotal),
            float(inv.tax),
            float(inv.total),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name='Arial', size=10)
            if fill:
                cell.fill = fill

    # Anchos
    for col, width in enumerate([8, 18, 28, 14, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="ventas_{date_from}_{date_to}.xlsx"'
    )
    wb.save(response)
    return response


@permission_required_any('billing.descargar_reportes_financieros')
def report_sales_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    import io

    date_from, date_to = _get_report_dates(request)
    invoices = list(_sales_queryset(date_from, date_to))
    summary  = _sales_summary(invoices)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    espresso = colors.HexColor('#231A10')
    rust     = colors.HexColor('#B5441B')

    elements = []
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 textColor=espresso, fontSize=16, spaceAfter=4)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 textColor=colors.grey, fontSize=10, spaceAfter=12)

    elements.append(Paragraph('Reporte de Ventas', title_style))
    elements.append(Paragraph(
        f'{date_from.strftime("%d/%m/%Y")} al {date_to.strftime("%d/%m/%Y")}', sub_style
    ))

    # Resumen
    summary_data = [
        ['Facturas', str(summary['count']),
         'Subtotal', f"${summary['subtotal']}",
         'IVA', f"${summary['tax']}",
         'Total', f"${summary['total']}"],
    ]
    t_sum = Table(summary_data, colWidths=[3*cm, 2*cm, 2.5*cm, 3*cm, 2*cm, 2.5*cm, 2.5*cm, 3*cm])
    t_sum.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), espresso),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [espresso]),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#F8F3EE')),
        ('TEXTCOLOR', (3, 0), (3, 0), rust),
        ('TEXTCOLOR', (5, 0), (5, 0), colors.HexColor('#F8F3EE')),
        ('TEXTCOLOR', (7, 0), (7, 0), rust),
    ]))
    elements.append(t_sum)
    elements.append(Spacer(1, 0.5*cm))

    # Detalle
    header = ['#', 'Fecha', 'Cliente', 'Cédula', 'Subtotal', 'IVA', 'Total']
    rows = [header]
    for inv in invoices:
        rows.append([
            str(inv.id),
            inv.invoice_date.strftime('%d/%m/%Y'),
            inv.customer.full_name,
            inv.customer.dni,
            f'${inv.subtotal}',
            f'${inv.tax}',
            f'${inv.total}',
        ])

    col_widths = [1.5*cm, 3*cm, 7*cm, 3.5*cm, 3*cm, 3*cm, 3*cm]
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    row_colors = [colors.HexColor('#F8F3EE'), colors.white]
    t.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), espresso),
        ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',       (2, 1), (2, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), row_colors),
        ('GRID',        (0, 0), (-1, -1), 0.3, colors.HexColor('#DDD3C5')),
        ('TEXTCOLOR',   (-1, 1), (-1, -1), rust),
        ('FONTNAME',    (-1, 1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="ventas_{date_from}_{date_to}.pdf"'
    )
    return response


@permission_required_any('billing.view_product', 'billing.descargar_reportes_financieros')
def report_stock(request):
    query = request.GET.get('q', '').strip()
    low_only = request.GET.get('low', '') == '1'
    threshold = 5

    products = Product.objects.filter(is_active=True).select_related('brand', 'group').order_by('stock')
    if query:
        products = products.filter(name__icontains=query)
    if low_only:
        products = products.filter(stock__lte=threshold)

    return render(request, 'billing/report_stock.html', {
        'products':  products,
        'query':     query,
        'low_only':  low_only,
        'threshold': threshold,
        'total_low': Product.objects.filter(is_active=True, stock__lte=threshold).count(),
    })


@permission_required_any('billing.descargar_reportes_financieros')
def report_stock_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import date

    query    = request.GET.get('q', '').strip()
    low_only = request.GET.get('low', '') == '1'
    threshold = 5

    products = Product.objects.filter(is_active=True).select_related('brand', 'group').order_by('stock')
    if query:
        products = products.filter(name__icontains=query)
    if low_only:
        products = products.filter(stock__lte=threshold)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte de Stock — {date.today().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13, name='Arial', color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='231A10')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Producto', 'Marca', 'Categoría', 'Precio', 'Stock', 'Estado']
    header_fill = PatternFill('solid', start_color='231A10')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    red_fill  = PatternFill('solid', start_color='FEF2F0')
    alt_fill  = PatternFill('solid', start_color='F8F3EE')
    for i, p in enumerate(products, 3):
        is_low = p.stock <= threshold
        row_fill = red_fill if is_low else (alt_fill if i % 2 == 0 else None)
        data = [p.name, p.brand.name, p.group.name, float(p.unit_price), p.stock,
                'Stock bajo' if p.stock == 0 else ('Poco stock' if is_low else 'OK')]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name='Arial', size=10,
                           color='C0392B' if is_low else '000000')
            if row_fill:
                cell.fill = row_fill

    for col, width in enumerate([30, 18, 18, 12, 10, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="stock_{date.today()}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
# Configuración del negocio
# ─────────────────────────────────────────────

@permission_required_any('billing.change_confignegocio')
def config_negocio_edit(request):
    """Vista para editar la configuración global del negocio."""
    from billing.models import ConfigNegocio
    from django.db import OperationalError, ProgrammingError
    try:
        config = ConfigNegocio.get()
    except (OperationalError, ProgrammingError):
        messages.error(request, 'La tabla de configuración no existe aún. Ejecuta las migraciones.')
        return redirect("billing:home")

    if request.method == 'POST':
        config.nombre_tienda    = (request.POST.get('nombre_tienda', '').strip() or config.nombre_tienda)[:80]
        config.slogan           = request.POST.get('slogan', '').strip()[:160]
        config.color_primario   = request.POST.get('color_primario', '#B5441B').strip()[:7]
        config.color_oscuro     = request.POST.get('color_oscuro', '#231A10').strip()[:7]
        config.color_fondo      = request.POST.get('color_fondo', '#F8F3EE').strip()[:7]
        config.color_navbar     = request.POST.get('color_navbar', '#231A10').strip()[:7]
        config.color_texto      = request.POST.get('color_texto', '#231A10').strip()[:7]
        config.hero_titulo      = request.POST.get('hero_titulo', '').strip()[:100]
        if 'hero_imagen' in request.FILES:
            config.hero_imagen  = request.FILES['hero_imagen']
        config.sobre_activo     = 'sobre_activo' in request.POST
        config.sobre_titulo     = request.POST.get('sobre_titulo', '').strip()
        config.sobre_texto      = request.POST.get('sobre_texto', '').strip()
        if 'sobre_imagen' in request.FILES:
            config.sobre_imagen = request.FILES['sobre_imagen']
        config.porque_activo    = 'porque_activo' in request.POST
        config.porque_titulo    = request.POST.get('porque_titulo', '').strip()
        config.porque_1_icono   = request.POST.get('porque_1_icono', '').strip()
        config.porque_1_titulo  = request.POST.get('porque_1_titulo', '').strip()
        config.porque_1_texto   = request.POST.get('porque_1_texto', '').strip()
        config.porque_2_icono   = request.POST.get('porque_2_icono', '').strip()
        config.porque_2_titulo  = request.POST.get('porque_2_titulo', '').strip()
        config.porque_2_texto   = request.POST.get('porque_2_texto', '').strip()
        config.porque_3_icono   = request.POST.get('porque_3_icono', '').strip()
        config.porque_3_titulo  = request.POST.get('porque_3_titulo', '').strip()
        config.porque_3_texto   = request.POST.get('porque_3_texto', '').strip()
        config.banner_activo    = 'banner_activo' in request.POST
        config.banner_titulo    = request.POST.get('banner_titulo', '').strip()
        config.banner_subtitulo = request.POST.get('banner_subtitulo', '').strip()
        config.banner_cta       = request.POST.get('banner_cta', '').strip()
        config.ruc              = request.POST.get('ruc', '').strip()
        config.email_contacto   = request.POST.get('email_contacto', '').strip()
        config.telefono         = request.POST.get('telefono', '').strip()
        config.whatsapp         = request.POST.get('whatsapp', '').strip()
        config.direccion        = request.POST.get('direccion', '').strip()
        config.facebook_url     = request.POST.get('facebook_url', '').strip()
        config.instagram_url    = request.POST.get('instagram_url', '').strip()
        config.tiktok_url       = request.POST.get('tiktok_url', '').strip()
        config.razon_social           = request.POST.get('razon_social', '').strip()[:300]
        config.nombre_comercial       = request.POST.get('nombre_comercial', '').strip()[:300]
        config.codigo_establecimiento = (request.POST.get('codigo_establecimiento', '001').strip() or '001')[:3].zfill(3)
        config.punto_emision          = (request.POST.get('punto_emision', '001').strip() or '001')[:3].zfill(3)
        config.ambiente_sri           = request.POST.get('ambiente_sri', '1').strip() if request.POST.get('ambiente_sri') in ('1', '2') else '1'
        config.obligado_contabilidad  = 'obligado_contabilidad' in request.POST
        config.contribuyente_especial = request.POST.get('contribuyente_especial', '').strip()[:10]

        if 'logo' in request.FILES:
            config.logo = request.FILES['logo']
        elif 'logo_clear' in request.POST:
            config.logo = None

        if 'hero_imagen_clear' in request.POST:
            config.hero_imagen = None

        if 'sobre_imagen_clear' in request.POST:
            config.sobre_imagen = None

        config.save()
        log_action(request, 'config_saved', 'ConfigNegocio', 1, 'Configuración del negocio guardada')
        messages.success(request, 'Configuración guardada correctamente.')
        return redirect('billing:config_negocio')

    iconos_disponibles = [
        ('bi-truck', 'Envío'),
        ('bi-shield-check', 'Garantía'),
        ('bi-headset', 'Soporte'),
        ('bi-star', 'Calidad'),
        ('bi-heart', 'Confianza'),
        ('bi-clock', 'Rapidez'),
        ('bi-gift', 'Ofertas'),
        ('bi-cash-coin', 'Precio'),
        ('bi-patch-check', 'Certif.'),
        ('bi-people', 'Equipo'),
        ('bi-shop', 'Tienda'),
        ('bi-geo-alt', 'Ubicación'),
        ('bi-whatsapp', 'WhatsApp'),
        ('bi-telephone', 'Teléfono'),
        ('bi-bag-check', 'Compra'),
        ('bi-award', 'Premio'),
        ('bi-lightning', 'Velocidad'),
        ('bi-box-seam', 'Producto'),
    ]
    porque_items = [
        (1, config.porque_1_icono, config.porque_1_titulo, config.porque_1_texto),
        (2, config.porque_2_icono, config.porque_2_titulo, config.porque_2_texto),
        (3, config.porque_3_icono, config.porque_3_titulo, config.porque_3_texto),
    ]
    return render(request, 'billing/config_negocio.html', {
        'config': config,
        'iconos_disponibles': iconos_disponibles,
        'porque_items': porque_items,
    })


# ─────────────────────────────────────────────
# Gestión de usuarios del panel
# ─────────────────────────────────────────────

@permission_required_any('auth.add_user', 'auth.change_user')
def user_management(request):
    from django.contrib.auth import get_user_model
    from django.contrib.auth.models import Group

    User = get_user_model()
    groups = Group.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'set_group':
            user_id = request.POST.get('user_id')
            group_name = request.POST.get('group_name', '')
            try:
                u = User.objects.get(pk=user_id)
                u.groups.clear()
                if group_name:
                    g = Group.objects.get(name=group_name)
                    u.groups.add(g)
                messages.success(request, f'Rol de {u.username} actualizado.')
            except (User.DoesNotExist, Group.DoesNotExist):
                messages.error(request, 'Usuario o grupo no encontrado.')

        elif action == 'create_user':
            username = request.POST.get('username', '').strip()
            email = request.POST.get('email', '').strip()
            group_name = request.POST.get('group_name', '')
            if not username:
                messages.error(request, 'El nombre de usuario es obligatorio.')
            elif User.objects.filter(username=username).exists():
                messages.error(request, f'El usuario "{username}" ya existe.')
            elif not email:
                messages.error(request, 'El correo electrónico es obligatorio para crear un usuario.')
            elif User.objects.filter(email=email).exists():
                messages.error(request, f'El correo "{email}" ya está en uso por otro usuario.')
            else:
                u = User.objects.create_user(username=username, email=email, is_active=False)
                u.set_unusable_password()
                u.save(update_fields=['password'])
                if group_name:
                    try:
                        u.groups.add(Group.objects.get(name=group_name))
                    except Group.DoesNotExist:
                        pass
                from billing.services import _send_panel_verification_code
                _send_panel_verification_code(u, request=request)
                messages.success(request, f'Usuario {username} creado. Se ha enviado un enlace de verificación a {email}.')

        return redirect('billing:user_management')

    from django.core.paginator import Paginator
    users_qs = User.objects.prefetch_related('groups').order_by('username')
    paginator = Paginator(users_qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'billing/user_management.html', {
        'users': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'groups': groups,
    })


@permission_required_any('auth.delete_user')
def delete_user(request, pk):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    target = get_object_or_404(User, pk=pk)

    if target.is_superuser:
        messages.error(request, 'No se puede eliminar un superusuario.')
        return redirect('billing:user_management')
    if target.pk == request.user.pk:
        messages.error(request, 'No puedes eliminar tu propia cuenta.')
        return redirect('billing:user_management')

    if request.method == 'POST':
        username = target.username
        target.delete()
        log_action(request, 'deleted', 'User', pk, f'Usuario eliminado: {username}')
        messages.success(request, f'Usuario {username} eliminado correctamente.')
        return redirect('billing:user_management')

    return render(request, 'billing/user_delete_confirm.html', {'target_user': target})


# ─────────────────────────────────────────────
# Verificación de código del panel
# ─────────────────────────────────────────────

def verify_panel_code(request):
    from django.contrib.auth import get_user_model
    from .models import PanelVerificationCode
    from billing.services import _send_panel_verification_code

    User = get_user_model()
    MAX_ATTEMPTS = 5
    LOCKOUT_MINUTES = 15

    if request.method == 'POST':
        action = request.POST.get('action', 'verify')

        if action == 'resend':
            email = request.POST.get('email', '').strip()
            if not email:
                messages.error(request, 'Ingresa tu correo.')
                return render(request, 'billing/verify_code.html')
            user = User.objects.filter(email=email).first()
            # Respuesta genérica para evitar enumeración de correos
            if user:
                _send_panel_verification_code(user, request=request)
            messages.success(request, 'Si el correo existe en el sistema, recibirás un nuevo código.')
            return render(request, 'billing/verify_code.html')

        # --- Protección contra fuerza bruta por sesión ---
        attempts = request.session.get('verify_attempts', 0)
        lockout_until = request.session.get('verify_lockout_until')
        if lockout_until:
            from django.utils import timezone as tz
            import datetime
            lockout_dt = tz.datetime.fromisoformat(lockout_until)
            if tz.now() < lockout_dt:
                remaining = int((lockout_dt - tz.now()).total_seconds() // 60) + 1
                messages.error(request, f'Demasiados intentos fallidos. Espera {remaining} minuto(s) antes de intentar de nuevo.')
                return render(request, 'billing/verify_code.html')
            else:
                request.session['verify_attempts'] = 0
                request.session['verify_lockout_until'] = None
                attempts = 0

        email = request.POST.get('email', '').strip()
        code = request.POST.get('code', '').strip()

        if not email or not code:
            messages.error(request, 'Completa todos los campos.')
            return render(request, 'billing/verify_code.html')

        import secrets as _secrets
        user = User.objects.filter(email=email).first()
        vc = PanelVerificationCode.objects.filter(user=user).order_by('-created_at').first() if user else None

        # Respuesta genérica para no revelar si el email existe (user enumeration)
        if not user or not vc:
            messages.error(request, 'Código incorrecto o expirado. Verifica los datos e intenta de nuevo.')
            return render(request, 'billing/verify_code.html')

        if vc.is_expired:
            messages.error(request, 'El código ha expirado. Solicita uno nuevo.')
            return render(request, 'billing/verify_code.html')

        # Comparación en tiempo constante para evitar timing attacks
        if not _secrets.compare_digest(vc.code, code):
            attempts += 1
            request.session['verify_attempts'] = attempts
            if attempts >= MAX_ATTEMPTS:
                from django.utils import timezone as tz
                import datetime
                lockout_until_dt = tz.now() + datetime.timedelta(minutes=LOCKOUT_MINUTES)
                request.session['verify_lockout_until'] = lockout_until_dt.isoformat()
                request.session['verify_attempts'] = 0
                messages.error(request, f'Demasiados intentos fallidos. Espera {LOCKOUT_MINUTES} minutos.')
            else:
                messages.error(request, f'Código incorrecto. Intentos restantes: {MAX_ATTEMPTS - attempts}.')
            return render(request, 'billing/verify_code.html')

        password = request.POST.get('password', '')
        if not password or len(password) < 8:
            messages.error(request, 'Crea una contraseña de al menos 8 caracteres.')
            return render(request, 'billing/verify_code.html')
        if password != request.POST.get('password_confirm', ''):
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'billing/verify_code.html')

        with transaction.atomic():
            user.set_password(password)
            user.is_active = True
            user.save()
            vc.delete()

        request.session.pop('verify_attempts', None)
        request.session.pop('verify_lockout_until', None)
        messages.success(request, f'Cuenta verificada. Ahora inicia sesión con tu usuario: {user.username} y la contraseña que creaste.')
        return redirect('login')

    verify_user = None
    email = request.GET.get('email', '')
    if email:
        verify_user = User.objects.filter(email=email).first()

    return render(request, 'billing/verify_code.html', {
        'verify_user': verify_user,
    })


# ─────────────────────────────────────────────
# Registro de actividad (Audit Log)
# ─────────────────────────────────────────────

@permission_required_any('billing.view_auditlog')
def activity_log(request):
    from billing.models import AuditLog
    from django.core.paginator import Paginator

    qs = AuditLog.objects.select_related('user')

    filter_user = request.GET.get('user', '').strip()
    filter_action = request.GET.get('action', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if filter_user:
        qs = qs.filter(user__username__icontains=filter_user)
    if filter_action:
        qs = qs.filter(action=filter_action)
    if date_from:
        qs = qs.filter(timestamp__date__gte=date_from)
    if date_to:
        qs = qs.filter(timestamp__date__lte=date_to)

    action_choices = AuditLog.ACTION_CHOICES
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    params = request.GET.copy()
    params.pop('page', None)

    return render(request, 'billing/activity_log.html', {
        'logs': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'search_params': params.urlencode(),
        'action_choices': action_choices,
        'filter_user': filter_user,
        'filter_action': filter_action,
        'date_from': date_from,
        'date_to': date_to,
    })


# ─────────────────────────────────────────────
# Promociones (envío masivo de correo)
# ─────────────────────────────────────────────

@permission_required_any('billing.change_customer')
def send_promotion(request):
    """Envía un correo HTML (promoción) a todos los clientes con email registrado."""
    from django.core.mail import EmailMultiAlternatives, get_connection
    from django.utils.html import strip_tags

    emails = list(
        Customer.objects.exclude(email__isnull=True).exclude(email='')
        .filter(accepts_promotions=True)
        .values_list('email', flat=True).distinct()
    )

    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        html_content = request.POST.get('html_content', '').strip()

        # Rate limiting: máximo 1 envío cada 30 minutos
        from billing.models import AuditLog
        import datetime
        from django.utils import timezone as _tz
        COOLDOWN = datetime.timedelta(minutes=30)
        last_promo = AuditLog.objects.filter(action='promotion_sent').order_by('-timestamp').first()
        if last_promo and (_tz.now() - last_promo.timestamp) < COOLDOWN:
            remaining = int((COOLDOWN - (_tz.now() - last_promo.timestamp)).total_seconds() // 60) + 1
            messages.error(request, f'Ya se envió una promoción recientemente. Espera {remaining} minuto(s) antes de volver a enviar.')
            return redirect('billing:send_promotion')

        if not subject or not html_content:
            messages.error(request, 'Completa el asunto y el contenido HTML antes de enviar.')
        elif not emails:
            messages.error(request, 'No hay clientes con correo registrado.')
        else:
            sent = 0
            connection = get_connection()
            connection.open()
            for email in emails:
                try:
                    message = EmailMultiAlternatives(
                        subject=subject,
                        body=strip_tags(html_content),
                        from_email=None,
                        to=[email],
                        connection=connection,
                    )
                    message.attach_alternative(html_content, 'text/html')
                    message.send()
                    sent += 1
                except Exception:
                    continue
            connection.close()
            log_action(request, 'promotion_sent', 'Customer', 0,
                       f'Promoción "{subject}" enviada a {sent} de {len(emails)} correos')
            messages.success(request, f'Promoción enviada a {sent} de {len(emails)} correos.')
        return redirect('billing:send_promotion')

    return render(request, 'billing/send_promotion.html', {'total_emails': len(emails)})
